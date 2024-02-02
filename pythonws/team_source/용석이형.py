from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image 
import time, re, os
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import tkinter as tk
from tkinter import ttk,Tk, Entry, Button, Radiobutton, StringVar, Frame, Label

options = Options()

# #최대 화면 조건
# options.add_argument('--start-maximized')
# # 화면 꺼짐 방지 조건
# options.add_experimental_option("detach", True)
# # 불필요한 에러메세지 제거 조건
# options.add_experimental_option("excludeSwitches", ["enable-logging"])

browser = webdriver.Chrome(options=options)

# headers = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'}

# 이전 스크롤 위치 초기화
last_height = browser.execute_script("return document.body.scrollHeight")



url = f'https://search.shopping.naver.com/book/home'
browser.get(url)
time.sleep(2)

actions = browser.find_element(By.TAG_NAME,'body')
# 무한 루프로 페이지의 끝까지 스크롤
while True:
    actions.send_keys(Keys.END) 
    time.sleep(0.5)
    # 현재 스크롤 위치 가져오기
    new_height = browser.execute_script("return document.body.scrollHeight")
    if new_height == last_height:
        break  # 스크롤이 더 이상 되지 않으면 루프 종료
    # 이전 스크롤 위치 업데이트
    last_height = new_height
    
#카테고리
elem = browser.find_element(By.CLASS_NAME, 'category_list_category__DqGyx')
categorylist = elem.text.split('\n')

root = Tk()
root.geometry()
root.title("데이터 입력")
root.option_add("*Font", "맑은고딕 12")
root.resizable(False, False)

list_label = Label(root, text='카테고리 목록입니다.')
list_label.pack()


get_url_exc_page = None
data_num_value = None
# Displaying the category list
selected_category = StringVar()
selected_category.set("")  # Set it to an empty string initially

def on_category_select(category):
    selected_category.set(category)

for category in categorylist:
    category_frame = Frame(root)
    category_frame.pack(anchor='w')
    category_radio = Radiobutton(category_frame, text=category, variable=selected_category, value=category) 
    category_radio.pack(side='left')
    category_label = Label(category_frame)
    category_label.pack(side='left')

ent_label = Label(root, text='데이터 갯수 입력:')
ent_label.pack()

ent = Entry(root)
ent.pack()

def btnpress():
    global get_url_exc_page, data_num_value ,selected_category_value
    # Accessing selected category and entered dataNum
    selected_category_value = selected_category.get()
    data_num_value = int(ent.get())
    
    print('Selected Category:', selected_category_value)
    print('Entered DataNum:', data_num_value)

    # Use the values as needed, for example, in your Selenium code
    xpath_idx = int(categorylist.index(selected_category_value)) + 1
    xpath = '//*[@id="container"]/div/div[10]/div/ul/li[{}]/a'.format(xpath_idx)
    elem = browser.find_element(By.XPATH, xpath)
    get_url_exc_page = elem.get_attribute('href')
    browser.execute_script('''
        var elements = document.querySelectorAll(".category_link_category__XlcyC");
        elements.forEach(function(element) {
            element.setAttribute("target", "_self");
        });
    ''')
    elem.click()
    time.sleep(1)

    root.destroy()

btn = Button(root)
btn.config(text="버튼", width=10, command=btnpress)
btn.pack()

root.mainloop()

print(data_num_value,type(data_num_value))

datalist = []

wb = Workbook()
ws = wb.active
ws.append(["제목","가격","발행일","순위"])


for i in range(1,(data_num_value//41)+2):
    getUrl = get_url_exc_page+'&pageIndex={}&pageSize=40'.format(i)
    browser.get(getUrl)
    time.sleep(2)
    actions = browser.find_element(By.TAG_NAME,'body')
    # 무한 루프로 페이지의 끝까지 스크롤
    while True:
        actions.send_keys(Keys.END) 
        time.sleep(0.5)
        # 현재 스크롤 위치 가져오기
        new_height = browser.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
             break  # 스크롤이 더 이상 되지 않으면 루프 종료
        # 이전 스크롤 위치 업데이트
        last_height = new_height

    price_elements = browser.find_elements(By.CSS_SELECTOR, '.bookPrice_price__zr5dh>em')
    name_elements = browser.find_elements(By.CSS_SELECTOR, '.bookListItem_title__X7f9_')
    published_date_elements = browser.find_elements(By.CLASS_NAME, 'bookListItem_detail_date___byvG')
    rank_elements = browser.find_elements(By.CSS_SELECTOR, '.bookListItem_feature__txTlp')

    for  name,price, date, score in zip(name_elements, price_elements, published_date_elements, rank_elements):
            
            published_date_text = date.get_attribute('textContent').strip()
            price_text = int(price.get_attribute('textContent').replace(',',''))
            name_text = name.get_attribute('textContent').strip()
            rank_text = score.get_attribute('textContent').strip()
        
            datalist.append(name_text)#가져온 문자열을 리스트에 추가(java는 push)
            ws.append([name_text ,price_text , published_date_text, rank_text]) 
           
            if len(datalist) >= data_num_value:
                break   


cleaned_category_name = re.sub('[/\\s]', '', selected_category_value) 
ws.title = "{}".format(cleaned_category_name)
efilename = '네이버북 {}_{}개.xlsx'.format(cleaned_category_name, data_num_value)

wb.save(efilename)


# 그래프 그리는 코드

plt.rcParams['font.family'] = 'Malgun Gothic'
plt.rcParams['axes.unicode_minus'] = False

data = pd.read_excel(efilename)

plt.rcParams['font.family'] = 'Malgun Gothic'
plt.rcParams['axes.unicode_minus'] = False

fig, axes = plt.subplots(1, 2, figsize=(12, 5))

price = data['가격'].sort_values(ascending=True)
axes[0].scatter(x=range(len(price)), y=price)
axes[0].set_xlabel('권 수')
axes[0].set_ylabel('가격')
axes[0].set_title('소설({}) 가격'.format(cleaned_category_name))

# Convert '발행일' to datetime format
data['발행일'] = pd.to_datetime(data['발행일'], errors='coerce')

# Extract only the year from the '발행일' column
data['연도'] = data['발행일'].dt.year

sns.countplot(x='연도', data=data, ax=axes[1])
axes[1].set_xlabel('연 도')
axes[1].set_ylabel('권 수')
axes[1].set_title('소설({}) 발행일'.format(cleaned_category_name))

plt.tight_layout()
plt.savefig('그래프.jpg',format='jpeg')

load_xlsx = load_workbook(efilename)
load_sheet = load_xlsx.active
img = Image('그래프.jpg')
load_sheet.add_image(img,'F1')
load_xlsx.save(efilename)

#메일보내는 코드
            
SMTP_SERVER = "smtp.naver.com"
SMTP_PORT = 587
SMTP_USER = "dudah789@naver.com" #네이버 메일주소
SMTP_PASSWORD = "dladudah123!" #네이버 메일비밀번호

recvs = "dudah789@naver.com" #받는 사람

msg =  MIMEMultipart()
msg["Subject"] = "제목" #제목
msg["From"] = SMTP_USER
msg["To"] = recvs

text = "테스트" #본문 텍스트

content = MIMEText(text)
msg.attach(content)

file_path = r'D:\zeromo\workspace\pythonws\{}'.format(efilename) #파일경로
with open(file_path,'rb') as f:
    attachment = MIMEApplication(f.read())
    attachment.add_header('Content-Disposition','attachment', filename = "{}".format(efilename)) #파일이름
    msg.attach(attachment)

s = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
s.starttls()
s.login(SMTP_USER, SMTP_PASSWORD)
s.sendmail(SMTP_USER, recvs, msg.as_string())
s.quit